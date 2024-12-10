from openpyxl import load_workbook

family_members_list = load_workbook('KAMAU FAMILY LIST.xlsx')
Sheet1 = family_members_list.active

wairumbi_family = [row[0]
                   for row in Sheet1.iter_rows(min_col=1, max_col=1, values_only=True) if row[0] is not None]
dorcas_family = [row[0]
                 for row in Sheet1.iter_rows(min_col=3, max_col=3, values_only=True) if row[0] is not None]
naomi_family = [row[0]
                for row in Sheet1.iter_rows(min_col=5, max_col=5, values_only=True) if row[0] is not None]
kungu_family = [row[0]
                for row in Sheet1.iter_rows(min_col=7, max_col=7, values_only=True) if row[0] is not None]
rahab_family = [row[0]
                for row in Sheet1.iter_rows(min_col=9, max_col=9, values_only=True) if row[0] is not None]
mary_family = [row[0]
               for row in Sheet1.iter_rows(min_col=11, max_col=11, values_only=True) if row[0] is not None]

all_families = {
    'wairumbi': wairumbi_family,
    'dorcas': dorcas_family,
    'naomi': naomi_family,
    'kungu': kungu_family,
    'rahab': rahab_family,
    'mary': mary_family,
}
