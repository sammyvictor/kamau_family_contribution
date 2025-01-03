import random

from datetime import date, timedelta
from openpyxl import Workbook, load_workbook


def shuffle_members(families_list):

    for family in families_list.values():
        random.shuffle(family)


def get_dates(start_year, start_month, start_day, end_year, end_month, end_day):
    start_date = date(start_year, start_month, start_day)
    end_date = date(end_year, end_month, end_day)
    current_date = start_date

    all_dates_to_assign = []
    while current_date <= end_date:
        all_dates_to_assign.append(current_date)
        current_date += timedelta(days=1)
    return all_dates_to_assign


# def leaders_assignment(dates_to_assign, families_list):
#     all_members = [{"name": member, "family": family}
#                    for family, members in families_list.items() for member in members]
#     timetable = {}
#     assigned_members = []
#     last_family_assigned = None

#     i = 0
#     while i < len(dates_to_assign):
#         current_date = dates_to_assign[i]
#         eligible_members = [
#             member for member in all_members
#             if member["name"] not in assigned_members and member["family"] != last_family_assigned
#         ]

#         day_of_week = current_date.isoweekday()

#         # Saturday handling (6) and Sunday handling (7)
#         if day_of_week == 6:  # Saturday
#             leader = random.choice(eligible_members)
#             timetable[current_date] = leader["name"]
#             # Assign same leader for Sunday
#             timetable[(current_date + timedelta(days=1))] = leader["name"]
#             assigned_members.append(leader["name"])
#             last_family_assigned = leader["family"]
#             i += 2  # Skip Sunday by advancing the index by 2
#         else:  # For weekdays
#             leader = random.choice(eligible_members)
#             timetable[current_date] = leader["name"]
#             assigned_members.append(leader["name"])
#             last_family_assigned = leader["family"]
#             i += 1

#         if i < len(dates_to_assign) - 1 and current_date.month != dates_to_assign[i + 1].month:
#             random.shuffle(all_members)
#             assigned_members.clear()
#     return timetable


def leaders_assignment(dates_to_assign, families_list):
    all_members = [{"name": member, "family": family}
                   for family, members in families_list.items() for member in members]
    timetable = {}
    assigned_members = set()
    last_family_assigned = None
    monthly_assigned = set()

    i = 0
    while i < len(dates_to_assign):
        current_date = dates_to_assign[i]
        day_of_week = current_date.isoweekday()

        # Eligible members: not assigned this month and not from the same family as the last leader
        eligible_members = [
            member for member in all_members
            if member["name"] not in monthly_assigned and member["family"] != last_family_assigned
        ]

        # If no eligible members are left for the month, reset monthly assignments
        if not eligible_members:
            monthly_assigned.clear()
            eligible_members = [
                member for member in all_members if member["family"] != last_family_assigned
            ]

        # Select a leader from eligible members
        leader = random.choice(eligible_members)
        timetable[current_date] = leader["name"]

        # Handle Saturday and Sunday assignment
        if day_of_week == 6:  # Saturday
            # Assign the same for Sunday
            timetable[current_date + timedelta(days=1)] = leader["name"]
            i += 2  # Skip Sunday by advancing index
        else:
            i += 1

        # Update tracking variables
        monthly_assigned.add(leader["name"])
        last_family_assigned = leader["family"]

        # Reset for a new month
        if i < len(dates_to_assign) and current_date.month != dates_to_assign[i].month:
            monthly_assigned.clear()
            assigned_members.clear()

    return timetable


def display_timetable(timetable):
    for date, leaders in timetable.items():
        # Adjusted for better formatting
        print(f"{date.strftime('%A, %Y-%m-%d')}: {''.join(leaders)}")


def save_timetable_to_excel(timetable, filename='kamau_family_contribution.xlsx'):
    kamau_wb = Workbook()
    tt_ws = kamau_wb.active
    tt_ws.title = 'Timetable'
    # Add headers
    tt_ws.append(["DAY", "DATE", "LEADER"])

    # Add the data
    for current_date, leader in timetable.items():
        formatted_day = current_date.strftime("%A")
        formatted_date = current_date.strftime("%d-%m-%Y")
        tt_ws.append([formatted_day, formatted_date, leader])
    kamau_wb.save(filename)


def rewrite_timetable_in_excel(timetable, filename='kamau_family_contribution.xlsx'):
    try:
        # Load the existing workbook
        kamau_wb = load_workbook(filename)
    except FileNotFoundError:
        # If the file doesn't exist, create a new workbook
        kamau_wb = Workbook()

    # Check if the 'Timetable' worksheet exists
    if 'Timetable' in kamau_wb.sheetnames:
        tt_ws = kamau_wb['Timetable']
        # Clear existing data
        for row in tt_ws.iter_rows(min_row=2, max_row=tt_ws.max_row, max_col=tt_ws.max_column):
            for cell in row:
                cell.value = None
    else:
        # Create a new worksheet if it doesn't exist
        tt_ws = kamau_wb.create_sheet('Timetable')

    # Add headers (overwrite if they already exist)
    tt_ws.delete_rows(1, 1)
    tt_ws.append(["DAY", "DATE", "LEADER"])

    # Add the data
    for current_date, leader in timetable.items():
        formatted_day = current_date.strftime("%A")
        formatted_date = current_date.strftime("%d-%m-%Y")
        tt_ws.append([formatted_day, formatted_date, leader])

    # Save the workbook (rewriting the file)
    kamau_wb.save(filename)
